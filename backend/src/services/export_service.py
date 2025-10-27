"""Export service for generating Excel and PDF reports.

Implements T203-T204 from tasks.md.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)
from sqlalchemy.orm import Session

from ..models.payment import Payment, PaymentStatus
from ..models.tenant import Tenant


class ExportService:
    """Service for exporting data to Excel and PDF formats."""

    def __init__(self, db: Session):
        self.db = db

    def export_payment_history_excel(
        self,
        tenant_id: UUID,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> BytesIO:
        """Export payment history to Excel format.

        Implements T203 from tasks.md.

        Args:
            tenant_id: Tenant UUID
            start_date: Start date for filtering
            end_date: End date for filtering

        Returns:
            BytesIO buffer containing Excel file
        """
        # Get tenant with related data
        tenant = self.db.query(Tenant).filter(Tenant.id == tenant_id).first()
        if not tenant:
            raise ValueError("Tenant not found")

        # Query payments
        query = self.db.query(Payment).filter(Payment.tenant_id == tenant_id)

        if start_date:
            query = query.filter(Payment.payment_date >= start_date)
        if end_date:
            query = query.filter(Payment.payment_date <= end_date)

        payments = query.order_by(Payment.payment_date.desc()).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment History"

        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = "Payment History Report"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        # Tenant info
        ws.merge_cells("A2:H2")
        info_cell = ws["A2"]
        info_cell.value = f"Tenant: {tenant.user.full_name if tenant.user else 'N/A'} | Property: {tenant.property.name if tenant.property else 'N/A'}"
        info_cell.alignment = Alignment(horizontal="center")

        # Date range
        if start_date or end_date:
            ws.merge_cells("A3:H3")
            date_cell = ws["A3"]
            date_range = f"Period: {start_date or 'Beginning'} to {end_date or 'Present'}"
            date_cell.value = date_range
            date_cell.alignment = Alignment(horizontal="center")
            header_row = 5
        else:
            header_row = 4

        # Headers
        headers = [
            "Date",
            "Amount",
            "Method",
            "Status",
            "Reference",
            "Gateway Fee",
            "Receipt",
            "Notes",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        data_start_row = header_row + 1
        total_amount = Decimal("0.00")

        for row_num, payment in enumerate(payments, data_start_row):
            ws.cell(row=row_num, column=1).value = (
                payment.payment_date.strftime("%Y-%m-%d") if payment.payment_date else ""
            )
            ws.cell(row=row_num, column=2).value = float(payment.amount)
            ws.cell(row=row_num, column=3).value = (
                payment.payment_method.value if payment.payment_method else ""
            )
            ws.cell(row=row_num, column=4).value = payment.status.value if payment.status else ""
            ws.cell(row=row_num, column=5).value = payment.transaction_reference or ""
            ws.cell(row=row_num, column=6).value = (
                float(payment.gateway_fee) if payment.gateway_fee else 0
            )
            ws.cell(row=row_num, column=7).value = payment.receipt_url or ""
            ws.cell(row=row_num, column=8).value = payment.notes or ""

            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).border = border_style

            # Sum total for confirmed payments
            if payment.status == PaymentStatus.CONFIRMED:
                total_amount += payment.amount

        # Total row
        total_row = data_start_row + len(payments) + 1
        ws.cell(row=total_row, column=1).value = "TOTAL PAID:"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2).value = float(total_amount)
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 30

        # Number formatting
        for row in range(data_start_row, total_row):
            ws.cell(row=row, column=2).number_format = "#,##0.00"
            ws.cell(row=row, column=6).number_format = "#,##0.00"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_payment_history_pdf(
        self,
        tenant_id: UUID,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> BytesIO:
        """Export payment history to PDF format.

        Implements T204 from tasks.md.

        Args:
            tenant_id: Tenant UUID
            start_date: Start date for filtering
            end_date: End date for filtering

        Returns:
            BytesIO buffer containing PDF file
        """
        # Get tenant with related data
        tenant = self.db.query(Tenant).filter(Tenant.id == tenant_id).first()
        if not tenant:
            raise ValueError("Tenant not found")

        # Query payments
        query = self.db.query(Payment).filter(Payment.tenant_id == tenant_id)

        if start_date:
            query = query.filter(Payment.payment_date >= start_date)
        if end_date:
            query = query.filter(Payment.payment_date <= end_date)

        payments = query.order_by(Payment.payment_date.desc()).all()

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1F4788"),
            spaceAfter=30,
            alignment=1,  # Center
        )

        styles["Heading2"]
        normal_style = styles["Normal"]

        # Title
        title = Paragraph("Payment History Report", title_style)
        elements.append(title)

        # Tenant info
        tenant_name = tenant.user.full_name if tenant.user else "N/A"
        property_name = tenant.property.name if tenant.property else "N/A"
        info_text = f"<b>Tenant:</b> {tenant_name}<br/><b>Property:</b> {property_name}"

        if start_date or end_date:
            date_range = (
                f"<br/><b>Period:</b> {start_date or 'Beginning'} to {end_date or 'Present'}"
            )
            info_text += date_range

        info_para = Paragraph(info_text, normal_style)
        elements.append(info_para)
        elements.append(Spacer(1, 0.3 * inch))

        # Payment table
        table_data = [["Date", "Amount", "Method", "Status", "Reference"]]

        total_amount = Decimal("0.00")

        for payment in payments:
            row = [
                payment.payment_date.strftime("%Y-%m-%d") if payment.payment_date else "",
                f"Rs. {float(payment.amount):,.2f}",
                payment.payment_method.value if payment.payment_method else "",
                payment.status.value if payment.status else "",
                payment.transaction_reference or "",
            ]
            table_data.append(row)

            if payment.status == PaymentStatus.CONFIRMED:
                total_amount += payment.amount

        # Total row
        table_data.append(["", f"Rs. {float(total_amount):,.2f}", "", "", "TOTAL PAID"])

        # Create table
        table = Table(
            table_data, colWidths=[1.2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.8 * inch]
        )

        table.setStyle(
            TableStyle(
                [
                    # Header row
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -2), 10),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),  # Amount column
                    ("GRID", (0, 0), (-1, -2), 1, colors.grey),
                    # Total row
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, -1), (-1, -1), 12),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                    ("GRID", (0, -1), (-1, -1), 1, colors.black),
                ]
            )
        )

        elements.append(table)

        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_text = f"<i>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
        footer_para = Paragraph(footer_text, normal_style)
        elements.append(footer_para)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def get_export_metadata(
        self,
        tenant_id: UUID,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> dict[str, Any]:
        """Get metadata for export (record count, date range, etc.).

        Args:
            tenant_id: Tenant UUID
            start_date: Start date
            end_date: End date

        Returns:
            Dict with export metadata
        """
        query = self.db.query(Payment).filter(Payment.tenant_id == tenant_id)

        if start_date:
            query = query.filter(Payment.payment_date >= start_date)
        if end_date:
            query = query.filter(Payment.payment_date <= end_date)

        payments = query.all()
        confirmed = [p for p in payments if p.status == PaymentStatus.CONFIRMED]

        total_amount = sum(p.amount for p in confirmed)

        return {
            "total_records": len(payments),
            "confirmed_payments": len(confirmed),
            "total_amount": float(total_amount),
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
        }
