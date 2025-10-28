"""Analytics API endpoints for financial data and insights.

Implements T096-T101, T109 from tasks.md.
"""

import logging
from datetime import date
from typing import Any
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession

from ...core.database import get_db
from ...core.security import get_current_user
from ...models.user import User, UserRole
from ...services.analytics_service import AnalyticsService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/analytics", tags=["analytics"])


@router.get("/rent-trends")
async def get_rent_collection_trends(
    start_date: date | None = Query(None, description="Start date for analysis"),
    end_date: date | None = Query(None, description="End date for analysis"),
    property_id: UUID | None = Query(None, description="Filter by property ID"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
) -> list[dict[str, Any]]:
    """Get rent collection trends over time.

    Returns monthly aggregated rent collection data including:
    - Total collected per month
    - Payment count
    - Completed vs pending amounts

    Accessible by: OWNER, INTERMEDIARY
    """
    # Check authorization
    if current_user.role not in [UserRole.OWNER, UserRole.INTERMEDIARY]:
        raise HTTPException(
            status_code=403, detail="Only owners and intermediaries can view analytics"
        )

    try:
        service = AnalyticsService(session)
        trends = await service.get_rent_collection_trends(
            user_id=current_user.id,
            property_id=property_id,
            start_date=start_date,
            end_date=end_date,
        )

        logger.info(
            f"User {current_user.id} retrieved rent collection trends "
            f"(property={property_id}, period={start_date} to {end_date})"
        )
        return trends

    except Exception as e:
        logger.error(f"Error retrieving rent collection trends: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to retrieve rent collection trends")


@router.get("/payment-status")
async def get_payment_status_overview(
    start_date: date | None = Query(None, description="Start date for analysis"),
    end_date: date | None = Query(None, description="End date for analysis"),
    property_id: UUID | None = Query(None, description="Filter by property ID"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    """Get payment status overview (completed, pending, failed).

    Returns breakdown of payment statuses with counts and amounts.

    Accessible by: OWNER, INTERMEDIARY
    """
    # Check authorization
    if current_user.role not in [UserRole.OWNER, UserRole.INTERMEDIARY]:
        raise HTTPException(
            status_code=403, detail="Only owners and intermediaries can view analytics"
        )

    try:
        service = AnalyticsService(session)
        overview = await service.get_payment_status_overview(
            user_id=current_user.id,
            property_id=property_id,
            start_date=start_date,
            end_date=end_date,
        )

        logger.info(
            f"User {current_user.id} retrieved payment status overview "
            f"(property={property_id}, period={start_date} to {end_date})"
        )
        return overview

    except Exception as e:
        logger.error(f"Error retrieving payment status overview: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to retrieve payment status overview")


@router.get("/expense-breakdown")
async def get_expense_breakdown(
    start_date: date | None = Query(None, description="Start date for analysis"),
    end_date: date | None = Query(None, description="End date for analysis"),
    property_id: UUID | None = Query(None, description="Filter by property ID"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
) -> list[dict[str, Any]]:
    """Get expense breakdown by bill type.

    Returns aggregated expense data grouped by bill type including:
    - Total amount per type
    - Bill count
    - Average amount
    - Percentage of total

    Accessible by: OWNER, INTERMEDIARY
    """
    # Check authorization
    if current_user.role not in [UserRole.OWNER, UserRole.INTERMEDIARY]:
        raise HTTPException(
            status_code=403, detail="Only owners and intermediaries can view analytics"
        )

    try:
        service = AnalyticsService(session)
        breakdown = await service.get_expense_breakdown(
            user_id=current_user.id,
            property_id=property_id,
            start_date=start_date,
            end_date=end_date,
        )

        logger.info(
            f"User {current_user.id} retrieved expense breakdown "
            f"(property={property_id}, period={start_date} to {end_date})"
        )
        return breakdown

    except Exception as e:
        logger.error(f"Error retrieving expense breakdown: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to retrieve expense breakdown")


@router.get("/revenue-expenses")
async def get_revenue_vs_expenses(
    start_date: date | None = Query(None, description="Start date for analysis"),
    end_date: date | None = Query(None, description="End date for analysis"),
    property_id: UUID | None = Query(None, description="Filter by property ID"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    """Get revenue vs expenses comparison.

    Returns:
    - Total revenue (completed rent payments)
    - Total expenses (bills)
    - Net profit
    - Profit margin percentage

    Accessible by: OWNER, INTERMEDIARY
    """
    # Check authorization
    if current_user.role not in [UserRole.OWNER, UserRole.INTERMEDIARY]:
        raise HTTPException(
            status_code=403, detail="Only owners and intermediaries can view analytics"
        )

    try:
        service = AnalyticsService(session)
        comparison = await service.get_revenue_vs_expenses(
            user_id=current_user.id,
            property_id=property_id,
            start_date=start_date,
            end_date=end_date,
        )

        logger.info(
            f"User {current_user.id} retrieved revenue vs expenses "
            f"(property={property_id}, period={start_date} to {end_date})"
        )
        return comparison

    except Exception as e:
        logger.error(f"Error retrieving revenue vs expenses: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to retrieve revenue vs expenses")


@router.get("/property-performance")
async def get_property_performance(
    start_date: date | None = Query(None, description="Start date for analysis"),
    end_date: date | None = Query(None, description="End date for analysis"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
) -> list[dict[str, Any]]:
    """Get performance comparison across all properties.

    Returns per-property data including:
    - Tenant count
    - Total revenue
    - Total expenses
    - Net profit
    - Occupancy rate

    Accessible by: OWNER, INTERMEDIARY
    """
    # Check authorization
    if current_user.role not in [UserRole.OWNER, UserRole.INTERMEDIARY]:
        raise HTTPException(
            status_code=403, detail="Only owners and intermediaries can view analytics"
        )

    try:
        service = AnalyticsService(session)
        performance = await service.get_property_performance(
            user_id=current_user.id,
            start_date=start_date,
            end_date=end_date,
        )

        logger.info(
            f"User {current_user.id} retrieved property performance "
            f"(period={start_date} to {end_date})"
        )
        return performance

    except Exception as e:
        logger.error(f"Error retrieving property performance: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to retrieve property performance")


@router.post("/export")
async def export_analytics_data(
    report_type: str = Query(
        ...,
        description="Type of report: rent-trends, payment-status, expense-breakdown, revenue-expenses, property-performance",
    ),
    format: str = Query("excel", description="Export format: excel or pdf"),
    start_date: date | None = Query(None, description="Start date for analysis"),
    end_date: date | None = Query(None, description="End date for analysis"),
    property_id: UUID | None = Query(None, description="Filter by property ID"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_db),
) -> dict[str, Any]:
    """Export analytics data to Excel or PDF.

    Generates a downloadable file with the requested analytics data.

    Accessible by: OWNER, INTERMEDIARY
    """
    # Check authorization
    if current_user.role not in [UserRole.OWNER, UserRole.INTERMEDIARY]:
        raise HTTPException(
            status_code=403, detail="Only owners and intermediaries can export analytics"
        )

    # Validate report type
    valid_report_types = [
        "rent-trends",
        "payment-status",
        "expense-breakdown",
        "revenue-expenses",
        "property-performance",
    ]
    if report_type not in valid_report_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid report type. Must be one of: {', '.join(valid_report_types)}",
        )

    # Validate format
    if format not in ["excel", "pdf"]:
        raise HTTPException(status_code=400, detail="Invalid format. Must be 'excel' or 'pdf'")

    try:
        service = AnalyticsService(session)

        # Get the appropriate data based on report type
        data = None
        if report_type == "rent-trends":
            data = await service.get_rent_collection_trends(
                user_id=current_user.id,
                property_id=property_id,
                start_date=start_date,
                end_date=end_date,
            )
        elif report_type == "payment-status":
            data = await service.get_payment_status_overview(
                user_id=current_user.id,
                property_id=property_id,
                start_date=start_date,
                end_date=end_date,
            )
        elif report_type == "expense-breakdown":
            data = await service.get_expense_breakdown(
                user_id=current_user.id,
                property_id=property_id,
                start_date=start_date,
                end_date=end_date,
            )
        elif report_type == "revenue-expenses":
            data = await service.get_revenue_vs_expenses(
                user_id=current_user.id,
                property_id=property_id,
                start_date=start_date,
                end_date=end_date,
            )
        elif report_type == "property-performance":
            data = await service.get_property_performance(
                user_id=current_user.id,
                start_date=start_date,
                end_date=end_date,
            )

        # Generate file based on format
        import tempfile
        import os
        from datetime import datetime as dt
        
        timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analytics_{report_type}_{timestamp}"
        
        if format == "xlsx":
            # Excel export using openpyxl
            file_path = await _generate_excel_export(data, report_type, filename)
        elif format == "pdf":
            # PDF export using reportlab
            file_path = await _generate_pdf_export(data, report_type, filename, start_date, end_date)
        else:
            # CSV export (simple format)
            file_path = await _generate_csv_export(data, report_type, filename)
        
        logger.info(
            f"User {current_user.id} exported {report_type} analytics "
            f"(format={format}, property={property_id}, period={start_date} to {end_date})"
        )

        return {
            "success": True,
            "report_type": report_type,
            "format": format,
            "file_path": file_path,
            "filename": os.path.basename(file_path),
            "data": data,
            "date_range": {
                "start": start_date.isoformat() if start_date else None,
                "end": end_date.isoformat() if end_date else None,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting analytics data: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to export analytics data")


# Helper functions for file export

async def _generate_excel_export(data: dict, report_type: str, filename: str) -> str:
    """Generate Excel file using openpyxl."""
    import tempfile
    import os
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        return await _generate_csv_export(data, report_type, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace("-", " ").title()
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Add title
    ws["A1"] = f"Analytics Report: {report_type.replace('-', ' ').title()}"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:D1")
    
    # Add data based on report type
    row = 3
    if isinstance(data, list):
        if len(data) > 0:
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header.replace("_", " ").title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for item in data:
                row += 1
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(key))
    else:
        # Handle dict data
        ws.cell(row=row, column=1, value="Metric").fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=2, value="Value").fill = header_fill
        ws.cell(row=row, column=2).font = header_font
        
        for key, value in data.items():
            row += 1
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=str(value))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, f"{filename}.xlsx")
    wb.save(file_path)
    
    logger.info(f"Generated Excel export: {file_path}")
    return file_path


async def _generate_pdf_export(
    data: dict, 
    report_type: str, 
    filename: str,
    start_date: date | None = None,
    end_date: date | None = None,
) -> str:
    """Generate PDF file using reportlab."""
    import tempfile
    import os
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        logger.warning("reportlab not installed, falling back to CSV")
        return await _generate_csv_export(data, report_type, filename)
    
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, f"{filename}.pdf")
    
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph(f"Analytics Report: {report_type.replace('-', ' ').title()}", title_style))
    
    # Date range
    if start_date and end_date:
        date_text = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Data table
    if isinstance(data, list) and len(data) > 0:
        headers = list(data[0].keys())
        table_data = [[h.replace("_", " ").title() for h in headers]]
        
        for item in data:
            row = [str(item.get(key, "")) for key in headers]
            table_data.append(row)
        
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
    else:
        # Handle dict data
        table_data = [["Metric", "Value"]]
        for key, value in data.items():
            table_data.append([key.replace("_", " ").title(), str(value)])
        
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
    
    doc.build(elements)
    logger.info(f"Generated PDF export: {file_path}")
    return file_path


async def _generate_csv_export(data: dict, report_type: str, filename: str) -> str:
    """Generate CSV file (fallback when libraries not available)."""
    import tempfile
    import os
    import csv
    
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, f"{filename}.csv")
    
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        if isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        else:
            writer = csv.writer(csvfile)
            writer.writerow(["Metric", "Value"])
            for key, value in data.items():
                writer.writerow([key, str(value)])
    
    logger.info(f"Generated CSV export: {file_path}")
    return file_path
